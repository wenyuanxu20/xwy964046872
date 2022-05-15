# coding = utf-8
import os
from openpyxl import Workbook
from openpyxl import load_workbook

# open
path = 'C:\\Users\\Wenyuan Xu\\Desktop\\excel-process' #\\Item_Config.xlsx'
#wb2 = load_workbook('test1.xlsx')


def get_all_cp_mapping_files(dir):
    file_list = []
    for root_dir, sub_dir, files in os.walk(r'' + dir):
        for file in files:
            # 此处可以按照要查找的文件名要求修改
            # 大概就是以什么为结尾，以什么为开头这样的判定条件，当然也可以用正则表达式。

            if file.endswith('.xlsx') and not file.startswith('~$'):
                file_name = os.path.join(root_dir, file)
                # 把拼接好的文件目录信息添加到列表中
                file_list.append(file_name)
    return file_list


file_list = get_all_cp_mapping_files(path)
print('all xlsx: ',file_list)

row_index = 0
target_row = []
for path in file_list:
    wb = load_workbook(path)
    target_row.append(path)
    for page in wb:  # 遍历单个excel文件每一页
        sheet_ranges = wb[page.title]
        #print('sheet ranges', sheet_ranges)
        for val in sheet_ranges.values:
            row_index += 1
            for target in val:
                target = str(target)
                if target == '30000001':  # 查找字符串目标
                    target_row.append(row_index)
        target_row.append(sheet_ranges)
        row_index = 0


print(target_row)


